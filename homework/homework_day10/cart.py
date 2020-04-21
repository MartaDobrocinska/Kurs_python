from openpyxl import load_workbook
from flask import Flask, render_template, request


class Db(object):
    def __init__(self):
        pass

    def get_items(self):
        """
        Function to get all items from db. Displays a list of products
        """
        data = load_workbook('db.xlsx')
        sheet = data.active
        self.products = {}
        for row in sheet.iter_rows(min_row=2,
                                   min_col=1,
                                   max_col=10,
                                   values_only=True):
            i_id = row[0]
            product = {
                "title": row[1],
                "author": row[6],
                "type": row[5],
                "pages": row[7],
                "quantity": row[4],
                "price": row[9]
            }
            self.products[i_id] = product
        return self.products

    def add_item(self):
        """
        Function to add an item to DB
        :param Book/Ebook object - object of book/ebook?

        """
        self.products.update()
        pass

    def remove_item(self, i_id):
        """
        Function to remove an item from DB
        :param i_id - id of item
        """
        self.products.pop(i_id)
        pass

    def update_item(self, key, value):
        """
        Function to update an item in DB
        :param Book/Ebook object - object of book/ebook?
        """
        self.products.update(key=value)
        pass


class Item(Db):
    def __init__(self, quantity=0, price=0.0, name='item', vat='23'):
        super().__init__()
        self.price = price
        self.name = name
        self.quantity = quantity
        self.vat = vat
        # self.created_at = created_at
        # self.last_buy_at = last_buy

    pass

    def get_item(self, i_id):
        """
        Function to get an item from db. Displays properties of a chosen product
        """
        items = Db()
        i_id = [1, 2, 3]
        item = items.get_items()[i_id]
        return item


    def price_gross(self):
        return self.price + self.price * (int(self.vat) / 100)


class Book(Item):
    vat = 23

    def __init__(self, quantity, title, author, pages, price):
        super().__init__()
        self.price = price
        self.author = author
        self.title = title
        self.pages = pages
        self.format = None
    pass


class Ebook(Book):
    vat = 8

    def __init__(self, quantity, title, author, pages, price):
        super().__init__(quantity, title, author, pages, price)
        self.vat = 8
        self.name = 'ebook'
        self.format = 'epub'

    pass


class Cart(Db):
    def __init__(self):
        super().__init__()
        self.elements = []
        self.quantity = 0
        self.nett = 0
        self.gross = 0

    def __str__(self):
        return str(self.elements)

    def add(self, item):
        self.elements.append(item)
        self.quantity +=1
        self.nett = self.nett + item.price
        self.gross = round(self.gross + item.price_gross(), 2)

    def __len__(self):
        return len(self.elements)

    def amount_nett(self):
        return self.nett

    def amount_gross(self):
        return self.gross


choice = Item().get_items()
book_1 = Book(choice[1].get('quantity'),
              choice[1].get('title'),
              choice[1].get('author'),
              choice[1].get('pages'),
              25.99)

book_2 = Book(choice[2].get('quantity'),
              choice[2].get('title'),
              choice[2].get('author'),
              choice[2].get('pages'),
              34.67)

ebook_1 = Ebook(choice[3].get('quantity'),
                choice[3].get('title'),
                choice[3].get('author'),
                choice[3].get('pages'),
                30.25)
cart = Cart()

app = Flask(__name__)
app.debug = True


@app.route("/", methods=['POST', 'GET'])
def main():
    return render_template('index.html')


@app.route("/product-list", methods=['POST', 'GET'])
def product_list():
    print(request.method)
    if request.method == 'POST':
        chosen = request.form['chosen']
        if chosen == '1':
            cart.add(book_1)
        elif chosen == '2':
            cart.add(book_2)
        elif chosen == '3':
            cart.add(ebook_1)
    return render_template('product_list.html', choice=choice)


@app.route("/cart", methods=['POST', 'GET'])
def cart_items():
    items = f'''Ilość w koszyku: {len(cart)}\n
                        Wartość netto: {cart.nett} PLN\n
                        Wartość brutto: {cart.gross} PLN'''
    items = items.replace('\n', '<br>')
    return render_template('cart.html', items=items)


app.run()

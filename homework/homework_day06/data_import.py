from openpyxl import load_workbook
from prettytable import PrettyTable


def data_import():
    x = PrettyTable()
    file = input('Podaj ścieżkę pliku Excel: ')
    data = load_workbook(file)
    sheet = data.worksheets[0]
    x.field_names = [cell.value for cell in sheet[1]]
    row_count = sheet.max_row
    for row in range(2, row_count+1):
        row = sheet[row]
        row_list = [row[x].value for x in range(len(row))]
        x.add_row(row_list)
    print(x)
